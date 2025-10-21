#!/usr/bin/env python3
"""
PTL Workshop Excel Data Importer - All-in-One

This script processes Excel files in the PTL_Workshop folder and imports
data from the "Database data" section into the report table.

Usage:
    python import_ptl_workshop.py                    # Interactive menu
    python import_ptl_workshop.py --analyze          # Analyze Excel files
    python import_ptl_workshop.py --dry-run          # Test import
    python import_ptl_workshop.py --import           # Import data
    python import_ptl_workshop.py --help             # Show help
"""

import os
import sys
import argparse
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime

import openpyxl
import requests
from dotenv import load_dotenv

# === Settings ===
BASE_DIR = Path(__file__).resolve().parent
PTL_FOLDER = BASE_DIR / "test_data" / "PTL_Workshop"

# Load environment variables
env_path = Path(__file__).resolve().parents[2] / 'frontend' / '.env'
load_dotenv(dotenv_path=env_path)

# API configuration
API_URL = os.getenv("VITE_API_BASE_URL")
API_REPORTS = f"{API_URL}/reports"

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('ptl_import.log')
    ]
)
logger = logging.getLogger(__name__)


class PTLWorkshopImporter:
    """PTL Workshop Excel data importer - All-in-One"""
    
    def __init__(self, folder_path: Path = PTL_FOLDER):
        self.folder_path = folder_path
        self.processed_files = []
        self.failed_files = []
        
    def find_database_data_row(self, worksheet) -> Optional[int]:
        """Find the row containing "Database data" in column A"""
        try:
            for row in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row, column=1).value
                if cell_value and isinstance(cell_value, str):
                    if "Database data" in cell_value:
                        logger.info(f"Found 'Database data' at row {row}")
                        return row
            return None
        except Exception as e:
            logger.error(f"Error searching for 'Database data': {e}")
            return None
    
    def extract_key_value_pairs(self, worksheet, start_row: int) -> Dict[str, str]:
        """Extract key-value pairs starting from the row after "Database data" """
        data = {}
        try:
            current_row = start_row + 1
            
            while current_row <= worksheet.max_row:
                key_cell = worksheet.cell(row=current_row, column=1)
                value_cell = worksheet.cell(row=current_row, column=2)
                
                if not key_cell.value:
                    break
                    
                key = str(key_cell.value).strip() if key_cell.value else ""
                value = str(value_cell.value).strip() if value_cell.value else ""
                
                if key:
                    data[key] = value
                    logger.debug(f"Extracted: {key} = {value}")
                
                current_row += 1
                
            logger.info(f"Extracted {len(data)} key-value pairs")
            return data
            
        except Exception as e:
            logger.error(f"Error extracting key-value pairs: {e}")
            return {}
    
    def map_data_to_report(self, data: Dict[str, str], filename: str) -> Dict[str, str]:
        """Map extracted data to Report model fields"""
        field_mapping = {
            # Basic info
            'op_name': 'op_name', 'Operator': 'op_name',
            'serial_num': 'serial_num', 'Serial Number': 'serial_num',
            'date': 'date', 'Date': 'date',
            
            # Platform info
            'os_version': 'os_version', 'OS Version': 'os_version',
            'platform_brand': 'platform_brand', 'Platform Brand': 'platform_brand',
            'platform': 'platform', 'Platform': 'platform',
            'platform_phase': 'platform_phase', 'Platform Phase': 'platform_phase',
            'platform_bios': 'platform_bios', 'Platform BIOS': 'platform_bios',
            'cpu': 'cpu', 'CPU': 'cpu',
            
            # Network
            'wlan': 'wlan', 'WLAN': 'wlan',
            'wlan_phase': 'wlan_phase', 'WLAN Phase': 'wlan_phase',
            'bt_driver': 'bt_driver', 'BT Driver': 'bt_driver',
            'bt_interface': 'bt_interface', 'BT Interface': 'bt_interface',
            'wifi_driver': 'wifi_driver', 'WiFi Driver': 'wifi_driver',
            'audio_driver': 'audio_driver', 'Audio Driver': 'audio_driver',
            
            # Software
            'wrt_version': 'wrt_version', 'WRT Version': 'wrt_version',
            'wrt_preset': 'wrt_preset', 'WRT Preset': 'wrt_preset',
            'msft_teams_version': 'msft_teams_version', 'Microsoft Teams Version': 'msft_teams_version',
            
            # Test scenario
            'scenario': 'scenario', 'Scenario': 'scenario',
            
            # Mouse
            'mouse_bt': 'mouse_bt', 'Mouse BT': 'mouse_bt',
            'mouse_brand': 'mouse_brand', 'Mouse Brand': 'mouse_brand',
            'mouse': 'mouse', 'Mouse': 'mouse',
            'mouse_click_period': 'mouse_click_period', 'Mouse Click Period': 'mouse_click_period',
            
            # Keyboard
            'keyboard_bt': 'keyboard_bt', 'Keyboard BT': 'keyboard_bt',
            'keyboard_brand': 'keyboard_brand', 'Keyboard Brand': 'keyboard_brand',
            'keyboard': 'keyboard', 'Keyboard': 'keyboard',
            'keyboard_click_period': 'keyboard_click_period', 'Keyboard Click Period': 'keyboard_click_period',
            
            # Headset
            'headset_bt': 'headset_bt', 'Headset BT': 'headset_bt',
            'headset_brand': 'headset_brand', 'Headset Brand': 'headset_brand',
            'headset': 'headset', 'Headset': 'headset',
            
            # Speaker
            'speaker_bt': 'speaker_bt', 'Speaker BT': 'speaker_bt',
            'speaker_brand': 'speaker_brand', 'Speaker Brand': 'speaker_brand',
            'speaker': 'speaker', 'Speaker': 'speaker',
            
            # Phone
            'phone_brand': 'phone_brand', 'Phone Brand': 'phone_brand',
            'phone': 'phone', 'Phone': 'phone',
            
            # Device1
            'device1_brand': 'device1_brand', 'Device1 Brand': 'device1_brand',
            'device1': 'device1', 'Device1': 'device1',
            
            # Power management
            'modern_standby': 'modern_standby', 'Modern Standby': 'modern_standby',
            'ms_period': 'ms_period', 'MS Period': 'ms_period',
            'ms_os_waiting_time': 'ms_os_waiting_time', 'MS OS Waiting Time': 'ms_os_waiting_time',
            
            's4': 's4', 'S4': 's4',
            's4_period': 's4_period', 'S4 Period': 's4_period',
            's4_os_waiting_time': 's4_os_waiting_time', 'S4 OS Waiting Time': 's4_os_waiting_time',
            
            'warm_boot': 'warm_boot', 'Warm Boot': 'warm_boot',
            'wb_period': 'wb_period', 'WB Period': 'wb_period',
            'wb_os_waiting_time': 'wb_os_waiting_time', 'WB OS Waiting Time': 'wb_os_waiting_time',
            
            'cold_boot': 'cold_boot', 'Cold Boot': 'cold_boot',
            'cb_period': 'cb_period', 'CB Period': 'cb_period',
            'cb_os_waiting_time': 'cb_os_waiting_time', 'CB OS Waiting Time': 'cb_os_waiting_time',
            
            # Other tests
            'microsoft_teams': 'microsoft_teams', 'Microsoft Teams': 'microsoft_teams',
            'apm': 'apm', 'APM': 'apm',
            'apm_period': 'apm_period', 'APM Period': 'apm_period',
            'opp': 'opp', 'OPP': 'opp',
            'swift_pair': 'swift_pair', 'Swift Pair': 'swift_pair',
            
            # Additional info
            'power_type': 'power_type', 'Power Type': 'power_type',
            'urgent_level': 'urgent_level', 'Urgent Level': 'urgent_level',
            'fix_work_week': 'fix_work_week', 'Fix Work Week': 'fix_work_week',
            'fix_bt_driver': 'fix_bt_driver', 'Fix BT Driver': 'fix_bt_driver',
            'jira_id': 'jira_id', 'JIRA ID': 'jira_id',
            'ips_id': 'ips_id', 'IPS ID': 'ips_id',
            'hsd_id': 'hsd_id', 'HSD ID': 'hsd_id',
            
            # Test results
            'result': 'result', 'Result': 'result',
            'fail_cycles': 'fail_cycles', 'Fail Cycles': 'fail_cycles',
            'cycles': 'cycles', 'Cycles': 'cycles',
            'duration': 'duration', 'Duration': 'duration',
            
            'log_path': 'log_path', 'Log Path': 'log_path',
        }
        
        mapped_data = {}
        
        # Map known fields
        for excel_key, report_field in field_mapping.items():
            if excel_key in data:
                mapped_data[report_field] = data[excel_key]
        
        # Set default values for required fields if not present
        required_defaults = {
            'op_name': 'Unknown',
            'serial_num': f'AUTO_{datetime.now().strftime("%Y%m%d_%H%M%S")}',
            'os_version': 'Unknown',
            'platform_brand': 'Unknown',
            'platform': 'Unknown',
            'platform_phase': 'Unknown',
            'platform_bios': 'Unknown',
            'cpu': 'Unknown',
            'wlan': 'Unknown',
            'wlan_phase': 'Unknown',
            'bt_driver': 'Unknown',
            'bt_interface': 'Unknown',
            'wifi_driver': 'Unknown',
            'audio_driver': 'Unknown',
            'wrt_version': 'Unknown',
            'wrt_preset': 'Unknown',
            'scenario': 'Unknown',
            'modern_standby': 'N/A',
            's4': 'N/A',
            'warm_boot': 'N/A',
            'cold_boot': 'N/A',
            'microsoft_teams': 'N/A',
            'apm': 'N/A',
            'opp': 'N/A',
            'swift_pair': 'N/A',
            'power_type': 'Unknown',
            'result': 'Unknown'
        }
        
        for field, default_value in required_defaults.items():
            if field not in mapped_data:
                mapped_data[field] = default_value
        
        # Add source file info
        mapped_data['log_path'] = f"PTL_Workshop/{filename}"
        
        logger.info(f"Mapped {len(mapped_data)} fields for report")
        return mapped_data
    
    def create_report_record(self, data: Dict[str, str]) -> bool:
        """Create a new report record via API"""
        try:
            response = requests.post(API_REPORTS, json=data, timeout=30)
            if response.status_code in (200, 201):
                logger.info(f"Successfully created report record via API")
                return True
            else:
                logger.error(f"API error creating report: {response.status_code}, {response.text}")
                return False
                
        except requests.exceptions.RequestException as e:
            logger.error(f"Request error creating report: {e}")
            return False
        except Exception as e:
            logger.error(f"Unexpected error creating report: {e}")
            return False
    
    def process_excel_file(self, file_path: Path, dry_run: bool = False) -> bool:
        """Process a single Excel file"""
        logger.info(f"Processing file: {file_path.name}")
        
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            # Find "Database data" row
            database_data_row = self.find_database_data_row(worksheet)
            if database_data_row is None:
                logger.warning(f"No 'Database data' section found in {file_path.name}")
                return False
            
            # Extract key-value pairs
            raw_data = self.extract_key_value_pairs(worksheet, database_data_row)
            if not raw_data:
                logger.warning(f"No data extracted from {file_path.name}")
                return False
            
            # Map to report fields
            mapped_data = self.map_data_to_report(raw_data, file_path.name)
            
            if dry_run:
                logger.info(f"DRY RUN: Would create report with data: {mapped_data}")
                return True
            
            # Create database record
            success = self.create_report_record(mapped_data)
            
            if success:
                self.processed_files.append(file_path.name)
            else:
                self.failed_files.append(file_path.name)
            
            return success
            
        except Exception as e:
            logger.error(f"Error processing {file_path.name}: {e}")
            self.failed_files.append(file_path.name)
            return False
    
    def process_folder(self, dry_run: bool = False) -> Tuple[int, int]:
        """Process all Excel files in the PTL_Workshop folder"""
        if not self.folder_path.exists():
            logger.error(f"Folder not found: {self.folder_path}")
            return 0, 0
        
        # Find all Excel files
        excel_files = list(self.folder_path.glob("*.xlsx")) + list(self.folder_path.glob("*.xls"))
        
        if not excel_files:
            logger.warning(f"No Excel files found in {self.folder_path}")
            return 0, 0
        
        logger.info(f"Found {len(excel_files)} Excel files to process")
        
        successful_count = 0
        failed_count = 0
        
        # Process each file
        for file_path in excel_files:
            try:
                if self.process_excel_file(file_path, dry_run):
                    successful_count += 1
                else:
                    failed_count += 1
            except Exception as e:
                logger.error(f"Unexpected error processing {file_path.name}: {e}")
                failed_count += 1
        
        return successful_count, failed_count
    
    def analyze_excel_files(self):
        """Analyze Excel files structure"""
        if not self.folder_path.exists():
            print(f"❌ PTL_Workshop folder not found: {self.folder_path}")
            return
        
        excel_files = list(self.folder_path.glob("*.xlsx")) + list(self.folder_path.glob("*.xls"))
        
        if not excel_files:
            print(f"❌ No Excel files found in {self.folder_path}")
            return
        
        print(f"🚀 Found {len(excel_files)} Excel file(s) to analyze")
        
        all_data = {}
        for file_path in excel_files:
            data = self._analyze_single_file(file_path)
            if data:
                all_data[file_path.name] = data
        
        # Summary across all files
        if all_data:
            print(f"\n{'='*60}")
            print(f"OVERALL SUMMARY")
            print(f"{'='*60}")
            
            all_keys = set()
            for file_data in all_data.values():
                all_keys.update(file_data.keys())
            
            print(f"📊 Total unique keys across all files: {len(all_keys)}")
            print(f"📁 Files processed: {len(all_data)}")
            
            if all_keys:
                print(f"\n🔑 All unique keys found:")
                for i, key in enumerate(sorted(all_keys), 1):
                    print(f"   {i:2d}. {key}")
    
    def _analyze_single_file(self, file_path: Path) -> Optional[Dict[str, str]]:
        """Analyze the structure of a single Excel file"""
        print(f"\n{'='*60}")
        print(f"ANALYZING: {file_path.name}")
        print(f"{'='*60}")
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            print(f"📊 Worksheet: {worksheet.title}")
            print(f"📏 Dimensions: {worksheet.max_row} rows × {worksheet.max_column} columns")
            
            # Search for "Database data" in column A
            database_data_row = None
            print(f"\n🔍 Searching for 'Database data' in column A...")
            
            for row in range(1, min(100, worksheet.max_row + 1)):
                cell_value = worksheet.cell(row=row, column=1).value
                if cell_value and isinstance(cell_value, str):
                    if "Database data" in cell_value:
                        database_data_row = row
                        print(f"🎯 Found 'Database data' at row {row}: '{cell_value}'")
                        break
                    elif row <= 30:
                        print(f"   Row {row:2d}: {cell_value}")
            
            if database_data_row is None:
                print("❌ 'Database data' not found in first 100 rows")
                return None
            
            # Extract data after "Database data"
            print(f"\n📋 Extracting key-value pairs starting from row {database_data_row + 1}:")
            print(f"{'Row':<4} | {'Key (Column A)':<30} | {'Value (Column B)'}")
            print(f"{'-'*4} | {'-'*30} | {'-'*20}")
            
            data_pairs = {}
            for row in range(database_data_row + 1, min(database_data_row + 51, worksheet.max_row + 1)):
                key_cell = worksheet.cell(row=row, column=1)
                value_cell = worksheet.cell(row=row, column=2)
                
                key = str(key_cell.value).strip() if key_cell.value else ""
                value = str(value_cell.value).strip() if value_cell.value else ""
                
                if not key:
                    print(f"{row:4d} | {'(empty - stopping)':<30} | ")
                    break
                
                data_pairs[key] = value
                print(f"{row:4d} | {key:<30} | {value}")
            
            # Count null/empty values
            null_count = sum(1 for value in data_pairs.values() if not value or value.strip() == "")
            
            print(f"\n📊 Summary:")
            print(f"   - Found {len(data_pairs)} key-value pairs")
            print(f"   - Found {null_count} keys with null/empty values")
            print(f"   - Data starts at row {database_data_row + 1}")
            
            return data_pairs
            
        except Exception as e:
            print(f"❌ Error analyzing file: {e}")
            return None
    
    def print_summary(self, successful_count: int, failed_count: int):
        """Print processing summary"""
        print("\n" + "="*50)
        print("PTL DATA IMPORT SUMMARY")
        print("="*50)
        print(f"Successfully processed: {successful_count} files")
        print(f"Failed to process: {failed_count} files")
        
        if self.processed_files:
            print(f"\nSuccessful files:")
            for filename in self.processed_files:
                print(f"  ✅ {filename}")
        
        if self.failed_files:
            print(f"\nFailed files:")
            for filename in self.failed_files:
                print(f"  ❌ {filename}")
        
        print("="*50)


def show_menu():
    """Show interactive menu"""
    print("\n" + "="*50)
    print("PTL WORKSHOP DATA IMPORT")
    print("="*50)
    print("1. Analyze Excel files (show structure)")
    print("2. Test import (dry run)")
    print("3. Import data to database")
    print("4. Exit")
    print("="*50)


def run_interactive():
    """Run interactive menu"""
    importer = PTLWorkshopImporter()
    
    while True:
        show_menu()
        
        try:
            choice = input("\nSelect an option (1-4): ").strip()
            
            if choice == '1':
                print("\n🔍 Analyzing Excel files...")
                importer.analyze_excel_files()
                
            elif choice == '2':
                print("\n🧪 Running test import (dry run)...")
                successful, failed = importer.process_folder(dry_run=True)
                importer.print_summary(successful, failed)
                
            elif choice == '3':
                print("\n⚠️  This will import data to the database!")
                confirm = input("Are you sure you want to continue? (y/N): ").strip().lower()
                
                if confirm == 'y':
                    print("\n📥 Importing data to database...")
                    successful, failed = importer.process_folder(dry_run=False)
                    importer.print_summary(successful, failed)
                else:
                    print("❌ Import cancelled")
                    
            elif choice == '4':
                print("👋 Goodbye!")
                break
            else:
                print("❌ Invalid choice. Please select 1-4.")
                
        except KeyboardInterrupt:
            print("\n\n👋 Goodbye!")
            break
        except Exception as e:
            print(f"❌ Unexpected error: {e}")
        
        input("\nPress Enter to continue...")


def main():
    """Main function"""
    parser = argparse.ArgumentParser(description="PTL Workshop Excel Data Importer")
    parser.add_argument("--analyze", action="store_true", help="Analyze Excel files structure")
    parser.add_argument("--dry-run", action="store_true", help="Test import without writing to database")
    parser.add_argument("--import", action="store_true", dest="do_import", help="Import data to database")
    parser.add_argument("--verbose", action="store_true", help="Enable verbose logging")
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    print(f"API_URL: {API_URL}")
    
    importer = PTLWorkshopImporter()
    
    if args.analyze:
        print("🔍 Analyzing Excel files...")
        importer.analyze_excel_files()
        
    elif args.dry_run:
        print("🧪 Running test import (dry run)...")
        successful, failed = importer.process_folder(dry_run=True)
        importer.print_summary(successful, failed)
        
    elif args.do_import:
        print("📥 Importing data to database...")
        successful, failed = importer.process_folder(dry_run=False)
        importer.print_summary(successful, failed)
        
    else:
        # Run interactive menu
        run_interactive()


if __name__ == "__main__":
    main()